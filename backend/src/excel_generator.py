
import logging
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.drawing.image import Image

from models import QAResults, ChallengeInput, ProblemStatement

logger = logging.getLogger(__name__)

class SolutionExcelGenerator:
    """Generate Excel solution files with embedded charts and SQL logic."""

    def __init__(self, output_path: Path):
        self.output_path = output_path

    def generate(self, qa_results: QAResults, problem_statement: ProblemStatement, data: Dict[str, pd.DataFrame]):
        """Generate the comprehensive solution Excel file."""
        logger.info("Generating Excel solution file...")
        
        wb = Workbook()
        
        # 1. Overview Sheet
        self._create_overview_sheet(wb, problem_statement)
        
        # 2. Question Sheets (One per question)
        for i, question in enumerate(problem_statement.analytical_questions, 1):
            sheet_name = f"Q{i} - Analysis"
            ws = wb.create_sheet(title=f"Q{i}")
            
            # Simulated analysis logic (in production this would come from an AI analyst agent)
            # For now, we generate template structure and placeholder data/charts
            self._create_question_sheet(ws, i, question, data)

        # Remove default sheet if empty
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        wb.save(self.output_path)
        logger.info(f"Excel solution saved to {self.output_path}")

    def _create_overview_sheet(self, wb: Workbook, problem: ProblemStatement):
        """Create the cover sheet with problem context."""
        ws = wb.active
        ws.title = "Overview"
        
        # Title
        ws["B2"] = problem.title
        ws["B2"].font = Font(size=20, bold=True, color="3B82F6")
        
        # Company Info
        ws["B4"] = f"Company: {problem.company_name}"
        ws["B5"] = f"Generated: {problem.generated_at.strftime('%Y-%m-%d')}"
        
        # Problem Statement
        ws["B7"] = "Problem Statement:"
        ws["B7"].font = Font(bold=True)
        ws["B8"] = problem.statement
        ws["B8"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells("B8:H20")
        
        # Disclaimer
        ws["B22"] = "This solution file contains analysis approach, SQL queries, and visual insights for each business question."
        ws["B22"].font = Font(italic=True, color="666666")

        # Column widths
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 60

    def _create_question_sheet(self, ws, q_num: int, question: str, data: Dict[str, pd.DataFrame]):
        """Create a detailed analysis sheet for a single question."""
        
        # Header
        ws["B2"] = f"Question {q_num}: {question}"
        ws["B2"].font = Font(size=14, bold=True, color="20C997")
        ws.merge_cells("B2:K2")
        
        # Sections
        sections = [
            ("Analysis Approach", 4, "Identify key metrics and group by relevant dimensions."),
            ("SQL Query Logic", 8, "SELECT dimension, SUM(metric) FROM table GROUP BY dimension"),
            ("Key Findings", 12, "• Trend indicates positive growth.\n• Segment A outperforms Segment B.")
        ]
        
        for title, row, content in sections:
            cell = ws.cell(row=row, column=2, value=title)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            
            content_cell = ws.cell(row=row+1, column=2, value=content)
            content_cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"B{row+1}:H{row+2}")

        # Data Table & Chart Area
        ws["B16"] = "Supporting Data"
        ws["B16"].font = Font(bold=True)
        
        # Generate dummy data for the chart since we don't have real query execution yet
        # Uses the first available numeric column from the first dataframe for demo purposes
        dummy_data = {
            "Category": ["A", "B", "C", "D"],
            "Value": [10, 25, 15, 30]
        }
        df = pd.DataFrame(dummy_data)
        
        # Write dataframe to sheet
        rows = dataframe_to_rows(df, index=False, header=True)
        start_row = 17
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=start_row + r_idx - 1, column=1 + c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.border = Border(bottom=Side(style="thin"))

        # Create Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Analysis Visualization"
        chart.y_axis.title = "Metric Value"
        chart.x_axis.title = "Dimension"

        data_ref = Reference(ws, min_col=3, min_row=17, max_row=17+len(df), max_col=3)
        cats_ref = Reference(ws, min_col=2, min_row=18, max_row=17+len(df))
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        ws.add_chart(chart, "F17") # Place chart next to data table
